#!/usr/bin/env python3
"""Merge Q1 extraction Markdown tables into one XLSX workbook."""

from __future__ import annotations

import re
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
MAIN_RUN = ROOT / "Evidence/question-extractions/8fc5df75-9595-43a7-bb4b-c078f94ea87a"
RETRY_RUN = ROOT / "Evidence/question-extractions/d3a1dcdf-933b-4a70-a8e3-67a09eef2d5d"
OUTPUT = (
    ROOT
    / "Evidence/question-extractions"
    / "Q1-evidence-all-rows-8fc5df75-plus-retry.xlsx"
)
EXCEL_CELL_LIMIT = 32767


def split_md_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [cell.strip() for cell in line.split("|")]


def is_separator(cells: list[str]) -> bool:
    if not cells:
        return False
    return all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) is not None for cell in cells)


def parse_q1_markdown(path: Path) -> tuple[list[str], list[list[str]]]:
    headers: list[str] = []
    rows: list[list[str]] = []
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        if not raw.strip().startswith("|"):
            continue
        cells = split_md_row(raw)
        if is_separator(cells):
            continue
        if not headers:
            headers = cells
            continue
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[: len(headers) - 1] + [" | ".join(cells[len(headers) - 1 :])]
        rows.append(cells)
    return headers, rows


def collect_markdown_files() -> dict[str, Path]:
    files: dict[str, Path] = {}
    for md in MAIN_RUN.glob("papers/*/Q1.md"):
        files[md.parent.name] = md
    if RETRY_RUN.is_dir():
        for md in RETRY_RUN.glob("papers/*/Q1.md"):
            files[md.parent.name] = md  # retry overrides failed main doc
    return files


def load_titles(document_ids: list[str]) -> dict[str, str]:
    if not document_ids:
        return {}
    values = ",".join(f"'{document_id}'" for document_id in document_ids)
    sql = (
        "SELECT document_id::text, coalesce(title, '') "
        f"FROM rag_document WHERE document_id IN ({values});"
    )
    result = subprocess.run(
        [
            "docker",
            "exec",
            "ai-code-postgres",
            "psql",
            "-U",
            "demo_01",
            "-d",
            "demo_01",
            "-t",
            "-A",
            "-F",
            "\t",
            "-c",
            sql,
        ],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    titles: dict[str, str] = {}
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        document_id, title = line.split("\t", 1)
        titles[document_id] = title
    return titles


def clip(value: str) -> str:
    if value is None:
        return ""
    text = str(value)
    if len(text) <= EXCEL_CELL_LIMIT:
        return text
    return text[: EXCEL_CELL_LIMIT - 12] + "\n[truncated]"


def main() -> None:
    md_by_doc = collect_markdown_files()
    titles = load_titles(sorted(md_by_doc))

    evidence_headers: list[str] | None = None
    all_rows: list[list[str]] = []
    docs_with_rows = 0
    empty_docs = 0

    for document_id in sorted(md_by_doc):
        headers, rows = parse_q1_markdown(md_by_doc[document_id])
        if not headers:
            empty_docs += 1
            continue
        if evidence_headers is None:
            evidence_headers = headers
        if not rows:
            empty_docs += 1
            continue
        docs_with_rows += 1
        title = titles.get(document_id, "")
        source_run = (
            "d3a1dcdf-933b-4a70-a8e3-67a09eef2d5d"
            if "d3a1dcdf-933b-4a70-a8e3-67a09eef2d5d" in str(md_by_doc[document_id])
            else "8fc5df75-9595-43a7-bb4b-c078f94ea87a"
        )
        for row_index, cells in enumerate(rows, start=1):
            all_rows.append(
                [document_id, title, source_run, str(row_index)]
                + [clip(cell) for cell in cells]
            )

    if evidence_headers is None:
        raise RuntimeError("No Q1 markdown headers found")

    meta_headers = ["document_id", "document_title", "extraction_run_id", "row_index"]
    headers = meta_headers + evidence_headers

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Q1证据行"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_idx, values in enumerate(all_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"

    widths = {
        1: 38,
        2: 42,
        3: 38,
        4: 10,
    }
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = widths.get(col, 24)

    # summary sheet
    summary = workbook.create_sheet("汇总", 0)
    summary.append(["item", "value"])
    summary.append(["main_run_id", "8fc5df75-9595-43a7-bb4b-c078f94ea87a"])
    summary.append(["retry_run_id", "d3a1dcdf-933b-4a70-a8e3-67a09eef2d5d"])
    summary.append(["markdown_documents", len(md_by_doc)])
    summary.append(["documents_with_rows", docs_with_rows])
    summary.append(["documents_empty_or_unparsed", empty_docs])
    summary.append(["evidence_rows", len(all_rows)])
    summary.append(["output", str(OUTPUT.relative_to(ROOT)).replace("\\", "/")])
    for col in (1, 2):
        summary.column_dimensions[get_column_letter(col)].width = 40

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"wrote {OUTPUT}")
    print(f"documents={len(md_by_doc)} with_rows={docs_with_rows} evidence_rows={len(all_rows)}")


if __name__ == "__main__":
    main()
